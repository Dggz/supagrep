"""Spreadsheet formatting"""
from contextlib import suppress
from typing import List

from cytoolz import concat
from openpyxl.cell import Cell
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.styles.colors import BLACK
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import Worksheet
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table

THIN_BLACK_SIDE = Side(border_style="thin", color=BLACK)
THIN_BLACK_BORDER = Border(
    top=THIN_BLACK_SIDE, left=THIN_BLACK_SIDE,
    right=THIN_BLACK_SIDE, bottom=THIN_BLACK_SIDE)
# header styles
SOLID_BLUE_FILL = PatternFill(patternType='solid', fgColor=Color('FF00B0F0'))
HEADER_FONT = Font(name='Arial', bold=True, italic=False, size=10)


def build_header(
        worksheet: Worksheet,
        headers: List[str],
        start_col: str = 'A') -> None:
    """Build worksheet header

    :param worksheet:
    :param headers:
    :param start_col:
    """
    for col_n, header in enumerate(headers, ord(start_col)):
        cell = worksheet['{}1'.format(column_format(col_n))]
        cell.value = header
        style_header_cell(cell)


def add_worksheet_table(
        worksheet: Worksheet,
        table_name: str,
        col: int,
        row: int,
        start_col: int = ord('A'),
        start_row: int = 1) -> None:
    """Adds table formatting to a worksheet

    Col and Row are the table limits

    :param worksheet:
    :param table_name:
    :param col:
    :param row:
    :param start_col:
    :param start_row:
    """
    tab = Table(
        displayName=table_name, ref='{}{}:{}{}'.format(
            column_format(start_col), start_row, column_format(col), row))
    worksheet.add_table(tab)


def set_cell_to_number(cell: Cell, cell_format: str = '0') -> Cell:
    """Sets the data type on a cell

    Sets the data type on a cell to number if the value of the cell is a number.
    Will not work on a list of numbers (cannot be processed as a single number).

    :param cell:
    :param cell_format:
    :return:
    """
    if cell.value and not cell.column == 'A':
        if cell.value.isdigit():
            cell.data_type = 'n'
            cell.number_format = cell_format
        with suppress(ValueError):
            if float(cell.value):
                cell.data_type = 'n'
                cell.number_format = cell_format
    return cell


def compute_column_dimensions(worksheet: Worksheet) -> None:
    """Provide good defaults for column dimensions

    :param worksheet:
    """
    for column_cells in worksheet.columns:
        cell_values = concat(
            [str(cell.value).split('\n') for cell in column_cells])
        length = max(len(cell_value or '') for cell_value in cell_values)
        if length > 150:
            worksheet.column_dimensions[
                column_cells[0].column].width = length + 5
        else:
            worksheet.column_dimensions[
                column_cells[0].column].width = (length + 2) * 1.2


def setup_auto_filter(worksheet: Worksheet, headers: List[str]) -> None:
    """Setup worksheet auto filter for those headers

    :param worksheet:
    :param headers:
    """
    worksheet.auto_filter = AutoFilter(
        ref='A1:{}1'.format(chr(ord('A') + len(headers) - 1)))


def style_header_cell(cell: Cell) -> None:
    """Stylize a header cell

    :param cell:
    """
    cell.fill = SOLID_BLUE_FILL
    cell.border = THIN_BLACK_BORDER
    cell.font = HEADER_FONT


def style_value_cell(cell: Cell) -> None:
    """Stylize a header cell

    :param cell:
    """
    cell.border = THIN_BLACK_BORDER


def column_format(col_n: int) -> str:
    """Formats the column number for Excel sheets

    Useful when you are populating/looking intro more than 26 columns ( > Z)
    The column names are composed of multiple characters (eg. AA, AB, CD)

    :param col_n:
    :return:
    """
    return get_column_letter(col_n - 64)
