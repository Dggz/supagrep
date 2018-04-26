"""XtremIO command entry point"""
import os

from openpyxl import load_workbook

from ntsparser.parsing import load_raw_content
from ntsparser.utils import get_bundle_dir, get_relevant_content, relevant_content_file_join
from ntsparser.xtremio.data_protection_groups import process as data_protection_groups
from ntsparser.xtremio.disks import process as disks
from ntsparser.xtremio.initiators_and_groups import process as initiators_and_groups
from ntsparser.xtremio.lun_mapping import process as lun_mapping
from ntsparser.xtremio.performance_output import process as performance_output
from ntsparser.xtremio.performance_summary import process as performance_summary
from ntsparser.xtremio.storage_array_summary import process as storage_array_summary
from ntsparser.xtremio.target_ports import process as target_ports
from ntsparser.xtremio.volume_performance import process as volume_performance
from ntsparser.xtremio.volumes import process as volumes
from ntsparser.xtremio.clusters import process as clusters_info

__all__ = ('main',)


def main(input_files: str, output_file: str) -> None:
    """XtremIO Entry point

    :param input_files:
    :param output_file:
    """
    template_path = os.path.join(
        get_bundle_dir(), r'resources\xtremio-template.xlsx')
    workbook = load_workbook(template_path)

    raw_content_patterns = (
        '*ShowClusters.out',
        '*ShowClustersInfo.out',
        '*ShowX-Bricks.out',
        '*ShowStorageControllersInfo.out',
        '*ShowClustersSavings.out',
        '*ShowSSDs.out',
        '*ShowTargets.out',
        '*ShowVolumes.out',
        '*ShowDataProtectionGroups.out',
        '*ShowVolumesPerformance.out',
        '*ShowLunMappings.out',
        '*ShowInitiators.out',
        '*ShowInitiatorGroups.out',
        '*performance_history.csv',
    )
    raw_content = load_raw_content(tuple(input_files), raw_content_patterns)
    storage_array_summary_content = get_relevant_content(
        raw_content, (raw_content_patterns[0],
                      raw_content_patterns[1],
                      raw_content_patterns[2],
                      raw_content_patterns[3],
                      raw_content_patterns[4]), '\n' + '*' * 20 + '\n')

    disks_content = get_relevant_content(
        raw_content, (raw_content_patterns[5], ), '*' * 20 + '\n')

    target_ports_content = get_relevant_content(
        raw_content, (raw_content_patterns[6], ), '*' * 20 + '\n')

    show_volumes_content = get_relevant_content(
        raw_content, (raw_content_patterns[7], ), '*' * 20 + '\n')

    data_protection_content = get_relevant_content(
        raw_content, (raw_content_patterns[8], ), '*' * 20 + '\n')

    volume_performance_content = get_relevant_content(
        raw_content, (raw_content_patterns[9], ), '*' * 20 + '\n')

    lun_mapping_content = get_relevant_content(
        raw_content, (raw_content_patterns[10], ), '*' * 20 + '\n')

    initiator_groups_content = get_relevant_content(
        raw_content, (raw_content_patterns[11],
                      raw_content_patterns[12]), '\n' + '*' * 20 + '\n')

    performance_content = get_relevant_content(
        raw_content, (raw_content_patterns[13], ), '\n' + '*' * 20 + '\n')

    performance_files = relevant_content_file_join(
        raw_content, (raw_content_patterns[13], ))

    clusters_files = relevant_content_file_join(
        raw_content, (raw_content_patterns[0], ))

    clusters = storage_array_summary(workbook, storage_array_summary_content)

    if len(performance_files) > 1:
        output_wb = output_file.split(os.sep)
        template_path = os.path.join(
            get_bundle_dir(), r'resources\xtremio-performance-template.xlsx')
        for file_content, file_cluster in \
                zip(performance_files, clusters_files):
            perf_workbook = load_workbook(template_path)
            output_wb[-1] = file_content.split('/')[0]
            perf_output = os.sep.join(output_wb) + '_perf.xlsx'
            file_content = '\n'.join(
                file for file in file_content.split('\n')[1:])
            perf_data = performance_output(perf_workbook, file_content)
            cluster_names = clusters_info(file_cluster)
            performance_summary(perf_workbook, perf_data + cluster_names)
            perf_workbook.save(perf_output)
    else:
        perf_data = performance_output(workbook, performance_content)
        performance_summary(workbook, perf_data + clusters)

    disks(workbook, disks_content)
    target_ports(workbook, target_ports_content)
    volumes(workbook, show_volumes_content)
    data_protection_groups(workbook, data_protection_content)
    volume_performance(workbook, volume_performance_content)
    lun_mapping(workbook, lun_mapping_content)
    initiators_and_groups(workbook, initiator_groups_content)

    workbook.save(output_file)
