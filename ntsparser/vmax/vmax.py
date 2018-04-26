"""VMAX command entry point"""
import os

from openpyxl import load_workbook

from ntsparser.parsing import load_raw_content
from ntsparser.utils import get_bundle_dir, get_relevant_content, relevant_content_file_join
from ntsparser.vmax.access_initiator import process as access_initiator
from ntsparser.vmax.access_view import process as access_view
from ntsparser.vmax.backend import process as backend
from ntsparser.vmax.device_name_list import process as device_name_list
from ntsparser.vmax.disks import process as disks
from ntsparser.vmax.dskgrp_summary import process as dskgrp_summary
from ntsparser.vmax.list_wwn import process as list_wwn
from ntsparser.vmax.requests import process as requests
from ntsparser.vmax.symcfg_list import process as symcfg_list
from ntsparser.vmax.symdev_info import process as symdev_info
from ntsparser.vmax.thin_devices import process as thin_devices

__all__ = ('main',)


def main(input_files: str, output_file: str) -> None:
    """VMAX Entry point

    :param input_files:
    :param output_file:
    """
    template_path = os.path.join(
        get_bundle_dir(), r'resources\vmax-template.xlsx')
    workbook = load_workbook(template_path)

    raw_content_patterns = (
        '*_symcfg_list.txt',
        '*_symaccess_list_devinfo.txt',
        '*_dskgrp_summary.txt',
        '*_symaccess_list_view.txt',
        '*_symaccess_show_init.txt',
        '*_list_tdev_gb_detail.txt',
        '*_symdev_list_identifier_device_name.txt',
        '*_symdev_list_wwn.txt',
        '*_type_backend.txt',
        '*_type_disks.txt',
        '*_type_requests.txt'
    )

    raw_content = load_raw_content(tuple(input_files), raw_content_patterns)
    symcfg_list_content = get_relevant_content(
        raw_content, (raw_content_patterns[0], ))

    symdev_info_content = get_relevant_content(
        raw_content, (raw_content_patterns[1], ), '*' * 20)

    dskgrp_summary_content = get_relevant_content(
        raw_content, (raw_content_patterns[2],))

    access_view_content = get_relevant_content(
        raw_content, (raw_content_patterns[3], ), '*' * 20)

    access_initiator_content = get_relevant_content(
        raw_content, (raw_content_patterns[4], ))

    thin_devices_content = get_relevant_content(
        raw_content, (raw_content_patterns[5], ))

    device_name_content = get_relevant_content(
        raw_content, (raw_content_patterns[6], ), '*' * 20)

    list_wwn_content = get_relevant_content(
        raw_content, (raw_content_patterns[7], ), '*' * 20)

    symcfg_list(workbook, symcfg_list_content)
    symdev_info(workbook, symdev_info_content)
    dskgrp_summary(workbook, dskgrp_summary_content)
    access_view(workbook, access_view_content)
    access_initiator(workbook, access_initiator_content)
    thin_devices(workbook, thin_devices_content)
    device_name_list(workbook, device_name_content)
    list_wwn(workbook, list_wwn_content)

    backend_content = relevant_content_file_join(
        raw_content, (raw_content_patterns[8], ))

    disks_content = relevant_content_file_join(
        raw_content, (raw_content_patterns[9], ))

    requests_content = relevant_content_file_join(
        raw_content, (raw_content_patterns[10], ))

    if len(backend_content) > 1:
        output_wb = output_file.split(os.sep)
        template_path = os.path.join(
            get_bundle_dir(), r'resources\vmax-performance-template.xlsx')
        for back_ct, disks_ct, requests_ct \
                in zip(backend_content, disks_content, requests_content):
            perf_workbook = load_workbook(template_path)
            output_wb[-1] = back_ct.split('_')[0]
            perf_output = os.sep.join(output_wb) + '_perf.xlsx'
            backend(perf_workbook, back_ct)
            disks(perf_workbook, disks_ct)
            requests(perf_workbook, requests_ct)

            perf_workbook.save(perf_output)
    else:
        backend(workbook, backend_content[0])
        disks(workbook, disks_content[0])
        requests(workbook, requests_content[0])

    workbook.save(output_file)
